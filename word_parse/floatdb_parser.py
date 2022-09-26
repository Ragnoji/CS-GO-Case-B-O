from selenium import webdriver
from openpyxl import Workbook
from time import sleep
from datetime import datetime

wb = Workbook()
ws = wb.active

starting_url = 'https://steamcommunity.com/openid/login?openid.claimed_id=http://specs.openid.net/auth/2.0/identifier_select&openid.identity=http://specs.openid.net/auth/2.0/identifier_select&openid.mode=checkid_setup&openid.ns=http://specs.openid.net/auth/2.0&openid.realm=https://csgofloat.com&openid.return_to=https://csgofloat.com/'

options = webdriver.ChromeOptions()
binary_yandex_driver_file = 'chromedriver.exe'
driver = webdriver.Chrome(binary_yandex_driver_file, options=options)
driver.get(starting_url)

_ = input('Press Enter when authorised')

ids_of_stickers = [5858, 5365]  # IDs of input stickers
ids_of_stickers = [str(n) for n in ids_of_stickers]
names_of_stickers = dict()

names_lines = open('current_names.txt', 'r', encoding='utf-8').readlines()

with open('current_items.txt', 'r', encoding='utf-8') as current_items:
    lines = current_items.readlines()
    for i, line in enumerate(lines[::-1]):
        if len(line.strip()) > 2 and line.strip()[1:-1] in ids_of_stickers:
            while "item_name" not in lines[-1 - i]:
                i -= 1
            item_name = lines[-1 - i][:lines[-1 - i].rfind('"')]
            item_name = item_name[item_name.rfind('"') + 2:]
            for line2 in names_lines[::-1]:
                if f'"{item_name}"' in line2:
                    item_name = line2[:line2.rfind('"')]
                    item_name = item_name[item_name.rfind('"') + 1:]
                    names_of_stickers[line.strip()[1:-1]] = item_name
                    break
print(names_of_stickers)

count_xpath = '/html/body/app-root/div/div[2]/app-float-db/div/div/app-float-dbtable/div/div/mat-card'
i = 1
for id in ids_of_stickers:
    address_n = f'A{i}'
    address_q = f'A{i + 1}'
    search_url = 'https://csgofloat.com/db?min=0&max=1&stickers='
    m = []
    for g in range(1, 5):
        driver.get(search_url + str([{'i': id} for _ in range(g)]).replace(' ', '').replace("'", '"'))
        while True:
            res = driver.find_elements_by_xpath(count_xpath)
            if res:
                break
            sleep(2)
        tmp = res[0].text.split()[1]
        if tmp.isdigit():
            tmp = int(tmp)
        m.append(tmp)
        sleep(2)
    for i in range(3):
        m[i] = m[i] - m[i + 1]
        m[i] = m[i] * (i + 1)
    m[3] = m[3] * 4
    quant = sum(m)
    ws[address_n] = names_of_stickers[id]
    ws[address_q] = quant
    i += 2

driver.close()
wb.save(f"{datetime.now().strftime('%Y-%m-%d')}.xlsx")
