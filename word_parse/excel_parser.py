from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
names = open('stickers.txt', 'r')
names = [line.strip() for line in names.readlines()]
ws.append(names)
tmp = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')

for index, name in enumerate(names):
    address = f'{tmp[index]}2'
    ws[address].hyperlink = f'https://steamcommunity.com/market/listings/730/Sticker | {name} | Antwerp 2022'
    ws[address].value = 'Paper'
    ws[address].style = "Hyperlink"

for index, name in enumerate(names):
    address = f'{tmp[index]}3'
    ws[address].hyperlink = f'https://steamcommunity.com/market/listings/730/Sticker | {name} (Holo) | Antwerp 2022'
    ws[address].value = 'Holo'
    ws[address].style = "Hyperlink"

for index, name in enumerate(names):
    address = f'{tmp[index]}4'
    ws[address].hyperlink = f'https://steamcommunity.com/market/listings/730/Sticker | {name} (Foil) | Antwerp 2022'
    ws[address].value = 'Foil'
    ws[address].style = "Hyperlink"

for index, name in enumerate(names):
    address = f'{tmp[index]}5'
    ws[address].hyperlink = f'https://steamcommunity.com/market/listings/730/Sticker | {name} (Gold) | Antwerp 2022'
    ws[address].value = 'Gold'
    ws[address].style = "Hyperlink"

# Save the file
wb.save("sample.xlsx")